import urllib.parse
import urllib.request
from urllib.error import HTTPError
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import time

nein = "Nein"
ja = "Ja"
error = "404 Fehler"

path="c:/test23.xlsx"
wb = load_workbook(path)
ws=wb.active
counter = 1

user_agent = 'Mozilla/5.0 (Windows; U; Windows NT 5.1; en-US; rv:1.9.0.7) Gecko/2009021910 Firefox/3.0.7'
headers={'User-Agent':user_agent,} 


for col in ws['A']:

    print(counter)   
    try:
        webpage = col.value
        request= urllib.request.Request(webpage,None,headers) #The assembled request
        response = urllib.request.urlopen(request)
        data = response.read() # The data u need
        output = data.decode('utf-8')
        soup = BeautifulSoup(output, 'html.parser')
        comparedata = soup.find_all('p')
        titel = soup.title.string
        if "beratung" in str(comparedata):
            ws.cell(row=counter, column=2).value = ja
            ws.cell(row=counter, column=4).value = titel
        elif "Beratung" in str(comparedata):
            ws.cell(row=counter, column=2).value = ja
            ws.cell(row=counter, column=4).value = titel  
        elif "beraten" in str(comparedata):
            ws.cell(row=counter, column=2).value = ja
            ws.cell(row=counter, column=4).value = titel
        elif "Beraten" in str(comparedata):
            ws.cell(row=counter, column=2).value = ja
            ws.cell(row=counter, column=4).value = titel
        else:
            ws.cell(row=counter, column=2).value = nein

        if "consulting" in str(comparedata):
            ws.cell(row=counter, column=3).value = ja
            ws.cell(row=counter, column=4).value = titel
        elif "Consulting" in str(comparedata):
            ws.cell(row=counter, column=3).value = ja
            ws.cell(row=counter, column=4).value = titel
        else:
            ws.cell(row=counter, column=3).value = nein  
        counter = counter + 1
        #wb.save("c:/test17.xlsx")
        time.sleep(2)            
    except HTTPError as err: 
        ws.cell(row=counter, column=4).value = error
        counter = counter + 1
        time.sleep(2)

wb.save("c:/test23.xlsx")
