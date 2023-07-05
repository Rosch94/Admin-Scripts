import re
from datetime import datetime
import os
from openpyxl import load_workbook

directory = 'C:\\MSXLOGS\W3SVC1'
path="c:/Dev/Test.xlsx"
wb = load_workbook(path)
ws=wb.active
counter = 1

for filename in os.listdir(directory):
    try:
        with open(os.path.join(directory, filename), 'r') as f:
            textlines = f.readlines()
            for line in textlines:
                match_pattern = re.search("^(?=.*\SSI)(?=.*/owa/ev.owa2).*$", line)
                if match_pattern is not None:
                    match_date = re.search(r'\d{4}-\d{2}-\d{2}', str(match_pattern))
                    date = datetime.strptime(match_date.group(), '%Y-%m-%d').date()

                    match_user = re.search(r"SSI\\\S+", line, re.IGNORECASE)
                    if match_user is not None:
                        ws.cell(row=counter, column=1).value = str(date)
                        ws.cell(row=counter, column=2).value = str(match_user.group())
                        counter = counter + 1
    except UnicodeDecodeError as err:
        print("Error")           
wb.save("c:/Dev/Test.xlsx")